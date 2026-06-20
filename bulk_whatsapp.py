"""
Fresh Go — Automatic WhatsApp Bulk Sender
==========================================
Double-click "FreshGo WhatsApp Sender.command" on Mac or
"FreshGo WhatsApp Sender.bat" on Windows.
"""

import time, sys, os, json, ssl, zipfile, shutil, subprocess
import urllib.request, urllib.parse
from datetime import datetime

try:
    import certifi
    _SSL_CTX = ssl.create_default_context(cafile=certifi.where())
except Exception:
    _SSL_CTX = ssl.create_default_context()

# ── Config ────────────────────────────────────────────────────────────────────
RAILWAY_URL  = "https://freshgo-production.up.railway.app"
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
SESSION_DIR  = os.path.join(BASE_DIR, "whatsapp_session")
CD_PATH      = os.path.join(BASE_DIR, "chromedriver.exe")   # Windows only

DEFAULT_MESSAGE = """Dear {name}, 🌿

📅 {date}

Your Fresh Go {product} ({qty}) has been delivered today. 100% pure and natural — straight from our Nankana Sahib farm to your home. 🐄

Thank you for choosing Fresh Go! ❤️
Any questions? Call/WhatsApp: 0300-3147887"""


# ── Phone helper ──────────────────────────────────────────────────────────────
def clean_phone(phone: str) -> str:
    phone = str(phone).strip().replace(" ", "").replace("-", "").replace("+", "")
    if phone.startswith("0"):
        phone = "92" + phone[1:]
    elif not phone.startswith("92"):
        phone = "92" + phone
    return phone


# ── ChromeDriver (Windows auto-download) ──────────────────────────────────────
def get_chromedriver_path() -> str:
    """Return path to chromedriver.exe, downloading if needed."""
    if os.path.exists(CD_PATH):
        return CD_PATH

    print("🔧 ChromeDriver download ho raha hai (sirf pehli baar)...")

    # Detect Chrome version from Windows registry
    major = "136"
    for reg_path in [
        r"HKLM\SOFTWARE\Google\Chrome\BLBeacon",
        r"HKCU\SOFTWARE\Google\Chrome\BLBeacon",
        r"HKLM\SOFTWARE\WOW6432Node\Google\Chrome\BLBeacon",
    ]:
        try:
            result = subprocess.run(
                ["reg", "query", reg_path, "/v", "version"],
                capture_output=True, text=True
            )
            import re
            m = re.search(r"(\d+)\.\d+\.\d+\.\d+", result.stdout)
            if m:
                major = m.group(1)
                break
        except Exception:
            pass

    print(f"   Chrome version detected: {major}.x")

    # Get exact ChromeDriver version for this Chrome major
    try:
        url = f"https://googlechromelabs.github.io/chrome-for-testing/LATEST_RELEASE_{major}"
        with urllib.request.urlopen(url, timeout=15, context=_SSL_CTX) as r:
            cd_version = r.read().decode().strip()
    except Exception:
        cd_version = f"{major}.0.7395.54"   # fallback guess

    zip_url = (
        f"https://storage.googleapis.com/chrome-for-testing-public"
        f"/{cd_version}/win64/chromedriver-win64.zip"
    )
    zip_path = os.path.join(BASE_DIR, "chromedriver.zip")

    print(f"   Downloading ChromeDriver {cd_version}...")
    urllib.request.urlretrieve(zip_url, zip_path)

    with zipfile.ZipFile(zip_path, "r") as zf:
        for name in zf.namelist():
            if name.endswith("chromedriver.exe"):
                zf.extract(name, BASE_DIR)
                extracted = os.path.join(BASE_DIR, name)
                shutil.move(extracted, CD_PATH)
                break

    os.remove(zip_path)
    extracted_dir = os.path.join(BASE_DIR, "chromedriver-win64")
    if os.path.isdir(extracted_dir):
        shutil.rmtree(extracted_dir)

    print("   ✅ ChromeDriver ready!\n")
    return CD_PATH


# ── Railway fetch ─────────────────────────────────────────────────────────────
def fetch_customers_from_railway() -> list:
    print("📡 Railway se customers fetch ho rahe hain...")
    try:
        url = f"{RAILWAY_URL}/api/todays-customers"
        with urllib.request.urlopen(url, timeout=15, context=_SSL_CTX) as resp:
            data = json.loads(resp.read().decode())
        customers = data.get("customers", [])
        for c in customers:
            if c.get("phone"):
                c["phone"] = clean_phone(str(c["phone"]))
        customers = [c for c in customers if c.get("phone")]
        print(f"✅ {len(customers)} customers mile Railway se\n")
        return customers
    except Exception as e:
        print(f"⚠️  Railway se fetch nahi hua: {e}")
        return []


# ── Excel reader ──────────────────────────────────────────────────────────────
def read_excel(filepath: str) -> list:
    try:
        import openpyxl, re
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        header_row = 1
        for r in range(1, 6):
            cells = [str(ws.cell(r, c).value or "").strip().lower()
                     for c in range(1, ws.max_column + 1)]
            if any("customer" in h or "phone" in h for h in cells):
                header_row = r
                break

        headers = [str(ws.cell(header_row, c).value or "").strip().lower()
                   for c in range(1, ws.max_column + 1)]

        def find_col(*keywords):
            for kw in keywords:
                for i, h in enumerate(headers):
                    if kw in h:
                        return i
            return None

        name_col    = find_col("customer", "name")
        phone_col   = find_col("phone", "number", "mobile")
        qty_col     = find_col("quantity", "qty", "litre")
        product_col = find_col("product")
        area_col    = find_col("area")
        rider_col   = find_col("rider")
        amount_col  = find_col("amount")
        payment_col = find_col("cash")

        if phone_col is None:
            print("❌ Excel mein Phone/Number column nahi mila!")
            return []

        is_delivery_report = rider_col is not None and area_col is not None
        customers = []

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(v is None for v in row):
                continue
            phone = str(row[phone_col] or "").strip()
            if not phone or phone.lower() == "none" or not any(ch.isdigit() for ch in phone):
                continue

            raw_name = str(row[name_col] or "Customer").strip() if name_col is not None else "Customer"
            name = re.sub(r"\s*\(\d+\)$", "", raw_name).strip() or "Customer"

            c = {
                "name":     name,
                "phone":    clean_phone(phone),
                "quantity": str(row[qty_col]     or "").strip() if qty_col     is not None else "",
                "product":  str(row[product_col] or "doodh").strip() if product_col is not None else "doodh",
            }
            if is_delivery_report:
                c["area"]    = str(row[area_col]   or "").strip()
                c["rider"]   = str(row[rider_col]  or "").strip()
                c["amount"]  = str(row[amount_col] or "").strip()
                c["payment"] = str(row[payment_col]or "").strip()
            customers.append(c)

        if is_delivery_report:
            print("ℹ️  Rider Delivery Report format detected")
        return customers
    except Exception as e:
        print(f"❌ Excel read error: {e}")
        return []


def find_excel_file() -> str | None:
    for f in sorted(os.listdir(BASE_DIR), reverse=True):
        if f.endswith((".xlsx", ".xls")) and not f.startswith("~"):
            return os.path.join(BASE_DIR, f)
    return None


def get_customers() -> list:
    customers = fetch_customers_from_railway()
    if customers:
        return customers

    print("ℹ️  Railway pe phone numbers nahi hain.")
    print("   Excel file se customers load kiye ja rahe hain...\n")

    excel_path = find_excel_file()
    if excel_path:
        print(f"📂 Excel mili: {os.path.basename(excel_path)}")
        customers = read_excel(excel_path)
        if customers:
            print(f"✅ {len(customers)} customers Excel se mile\n")
            return customers

    print("📂 Excel file ka path paste karo (ya Enter dabaiye cancel ke liye):")
    path = input("Path: ").strip().strip('"').strip("'")
    if path and os.path.exists(path):
        customers = read_excel(path)
        if customers:
            print(f"✅ {len(customers)} customers mile\n")
    return customers


# ── Message builder ───────────────────────────────────────────────────────────
def build_message(c: dict, message_template: str) -> str:
    name     = c.get("name", "Customer")
    quantity = c.get("quantity", "")
    product  = c.get("product", "milk")
    qty_text = f"{quantity} {product}".strip() if quantity else product
    unit     = "kg" if "ghee" in product.lower() else "L"
    qty_with_unit = f"{quantity} {unit}" if quantity else f"1 {unit}"
    d = datetime.now()
    today = f"{d.day} {d.strftime('%B %Y')}"

    area    = c.get("area", "")
    rider   = c.get("rider", "")
    amount  = c.get("amount", "")
    payment = c.get("payment", "")

    if rider and area and message_template == DEFAULT_MESSAGE:
        return (
            f"Dear {name}, 🌿\n\n"
            f"📅 {today}\n\n"
            f"Aaj aapki delivery complete ho gayi! ✅\n\n"
            f"📦 {product} — {qty_with_unit}\n"
            f"📍 Area: {area}\n"
            f"🚴 Rider: {rider}\n"
            + (f"💰 Amount: Rs. {amount}\n" if amount else "")
            + (f"💳 Payment: {payment}\n" if payment else "")
            + f"\nShukriya Fresh Go choose karne ke liye! ❤️\n"
            f"Mona Dairy Farms, Nankana Sahib 🐄\n"
            f"Call/WhatsApp: 0300-3147887"
        )
    return message_template.format(
        name=name, product=product,
        qty=qty_with_unit, quantity=qty_text, date=today,
    )


# ── Send via webbrowser + pyautogui (no ChromeDriver needed) ─────────────────
def send_messages(customers: list, message_template: str):
    import webbrowser
    try:
        import pyautogui
    except ImportError:
        print("❌ pyautogui install nahi. Bat file se dobara chalao.")
        sys.exit(1)

    print(f"🐄 Fresh Go WhatsApp Bulk Sender")
    print(f"📋 {len(customers)} customers ko message bheja jayega")
    print(f"\n⚠️  Sending shuru hogi — mouse mat hilao aur koi button mat dabao!\n")
    print("3 second mein shuru hogi...")
    time.sleep(3)

    def focus_chrome():
        """Bring Chrome window to front on Windows."""
        if sys.platform == "win32":
            try:
                subprocess.run(
                    ["powershell", "-Command",
                     "$wshell = New-Object -ComObject wscript.shell; "
                     "$wshell.AppActivate('Chrome')"],
                    capture_output=True
                )
                time.sleep(0.5)
            except Exception:
                pass

    sent = 0
    for i, c in enumerate(customers):
        name  = c.get("name", "Customer")
        phone = c["phone"]
        msg   = build_message(c, message_template)
        url   = (f"https://web.whatsapp.com/send?phone={phone}"
                 f"&text={urllib.parse.quote(msg)}")

        print(f"📤 [{i+1}/{len(customers)}] {name} ({phone})...")
        webbrowser.open(url)
        time.sleep(8)       # page load wait
        focus_chrome()      # Chrome ko foreground mein lao
        pyautogui.press('enter')   # message send
        time.sleep(3)
        sent += 1
        print(f"   ✅ Sent!")

    print(f"\n{'='*40}")
    print(f"✅ {sent} messages bhej diye!")
    print(f"{'='*40}")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    customers = get_customers()

    if not customers:
        print("Aaj koi customers nahi hain jinka phone number ho.")
        input("Enter dabaiye band karne ke liye...")
        sys.exit(0)

    print("Customers:")
    for c in customers:
        print(f"  • {c.get('name','?')} — {c['phone']} — {c.get('quantity','')} {c.get('product','')}")

    print(f"\nMessage template (Enter dabaiye default ke liye, ya apna likhein):")
    custom = input("Message: ").strip()
    template = custom if custom else DEFAULT_MESSAGE

    confirm = input(f"\n{len(customers)} customers ko message bhejna hai? (yes/no): ").strip().lower()
    if confirm in ("yes", "y", "ha", "haan"):
        send_messages(customers, template)
    else:
        print("Cancelled.")

    input("\nEnter dabaiye band karne ke liye...")
