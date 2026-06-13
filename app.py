import os
import asyncio
import secrets
import hmac
import json
import base64
from datetime import datetime, timedelta
from fastapi import FastAPI, Request, Query, UploadFile, File
from fastapi.responses import JSONResponse, HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv

from database       import init_db, get_all_orders, get_stats, update_order_status, \
                           get_all_scheduled_posts, delete_scheduled_post, create_scheduled_post, \
                           get_order, get_todays_customers, \
                           get_all_admin_users, get_admin_user_by_username, \
                           create_admin_user, delete_admin_user, update_admin_user, \
                           seed_staff_from_env
from ai_engine      import get_support_reply, generate_post
from facebook       import send_message, publish_post
from state          import set_pending_post, get_pending_post, clear_pending_post, has_pending_post
from rate_limiter   import is_allowed, seconds_until_reset
from order_flow     import (has_active_order, is_order_intent, is_history_intent,
                             start_order, handle_order_step, get_customer_order_history)
from notifications  import notify_customer_status
from scheduler_runner import start_scheduler, parse_schedule_command

load_dotenv()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = FastAPI(title="Fresh Go AI Agent")
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")

VERIFY_TOKEN   = os.getenv("WEBHOOK_VERIFY_TOKEN", "freshgo_secret_123")
ADMIN_FB_ID    = os.getenv("ADMIN_FB_ID", "")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "freshgo123")

SESSION_COOKIE   = "freshgo_session"
SESSION_HOURS    = 72   # 3 days — survives Railway restarts
# Secret used to sign session cookies — set SESSION_SECRET in Railway env for security
_SESSION_SECRET  = os.getenv("SESSION_SECRET", "freshgo_default_secret_change_me")

_PROTECTED_PREFIXES = ("/dashboard", "/api/")
_PUBLIC_API_PATHS   = {"/api/login", "/api/logout", "/api/todays-customers", "/health"}
_PUBLIC_PREFIXES    = ("/static/", "/webhook", "/whatsapp", "/favicon.ico")


def _sign(data: str) -> str:
    """HMAC-SHA256 sign a string."""
    import hashlib
    return hmac.new(_SESSION_SECRET.encode(), data.encode(), hashlib.sha256).hexdigest()


def _make_session_cookie(username: str, role: str) -> str:
    """Encode {username, role, expiry} + HMAC signature into a cookie value."""
    expiry = (datetime.utcnow() + timedelta(hours=SESSION_HOURS)).isoformat()
    payload = base64.urlsafe_b64encode(
        json.dumps({"u": username, "r": role, "e": expiry}).encode()
    ).decode()
    sig = _sign(payload)
    return f"{payload}.{sig}"


def _decode_session_cookie(cookie: str) -> dict | None:
    """Verify and decode a session cookie. Returns {username, role} or None."""
    try:
        payload, sig = cookie.rsplit(".", 1)
        if not hmac.compare_digest(sig, _sign(payload)):
            return None
        data = json.loads(base64.urlsafe_b64decode(payload.encode()).decode())
        if datetime.fromisoformat(data["e"]) < datetime.utcnow():
            return None
        return {"username": data["u"], "role": data["r"]}
    except Exception:
        return None


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path

    if path in {"/", "/login", "/health"} or \
       any(path.startswith(p) for p in _PUBLIC_PREFIXES) or \
       path in _PUBLIC_API_PATHS:
        return await call_next(request)

    if not any(path.startswith(p) for p in _PROTECTED_PREFIXES):
        return await call_next(request)

    cookie = request.cookies.get(SESSION_COOKIE)
    user   = _decode_session_cookie(cookie) if cookie else None
    if user:
        request.state.user = user
        return await call_next(request)

    if path.startswith("/api/"):
        return JSONResponse(status_code=401, content={"error": "Not authenticated"})
    return RedirectResponse(url=f"/login?next={path}", status_code=302)


CONFIRM_WORDS = {"yes", "yep", "post it", "confirm", "ok", "okay", "ha", "haan",
                 "kar do", "post", "send"}
CANCEL_WORDS  = {"no", "nahi", "cancel", "nope", "stop", "na"}


@app.on_event("startup")
async def startup():
    init_db()
    seed_staff_from_env()   # Creates/updates staff accounts from Railway env vars
    # Restore brand assets from DB in case the filesystem was reset (Railway redeploy)
    from brand_assets import get_asset_urls
    assets = get_asset_urls()
    if assets.get("logo_url"):
        print(f"[Startup] Logo ready: {assets['logo_url']}")
    if assets.get("packet_url"):
        print(f"[Startup] Packet ready: {assets['packet_url']}")
    start_scheduler()
    os.makedirs("static/images", exist_ok=True)
    os.makedirs("static/videos", exist_ok=True)
    print("🐄 Fresh Go AI Agent started!")


def is_admin(sender_id: str) -> bool:
    return bool(ADMIN_FB_ID) and sender_id == ADMIN_FB_ID


# ─────────────────────────────────────────────────────────────────────────────
# Admin flow
# ─────────────────────────────────────────────────────────────────────────────
async def process_admin(sender_id: str, text: str, local_mode: bool = False) -> list:
    replies    = []
    text_lower = text.strip().lower()

    # ── Pending post confirmation ─────────────────────────────────────────────
    if has_pending_post(sender_id):
        if any(w in text_lower for w in CONFIRM_WORDS):
            pending  = get_pending_post(sender_id)
            p_text   = pending["text"]
            img_url  = pending.get("image_url")
            vid_url  = pending.get("video_url")
            platform = pending.get("platform", "facebook")
            clear_pending_post(sender_id)

            if local_mode:
                replies.append({"type": "published", "text": p_text,
                                 "image_url": img_url, "video_url": vid_url})
                replies.append({"type": "message",
                                 "text": "✅ [Local Test] Post simulated! Connect accounts to go live. 🎉"})
            else:
                await _do_publish(replies, p_text, img_url, vid_url, platform)
            return replies

        elif any(w in text_lower for w in CANCEL_WORDS):
            clear_pending_post(sender_id)
            replies.append({"type": "message",
                             "text": "❌ Post cancelled. Send !post <topic> anytime."})
            return replies
        else:
            clear_pending_post(sender_id)
            replies.append({"type": "message", "text": "🔄 Regenerating with new topic..."})

    # ── !post <topic> ─────────────────────────────────────────────────────────
    if text_lower.startswith("!post "):
        topic = text[6:].strip()
        if not topic:
            replies.append({"type": "message",
                             "text": "Example: !post Eid discount on fresh milk 🎉"})
            return replies
        replies.append({"type": "message",
                         "text": f'⏳ Generating image post about "{topic}"...'})
        result = await generate_post(topic)
        set_pending_post(sender_id, {
            "text": result["text"], "image_url": result["image_url"],
            "video_url": None, "platform": "both"
        })
        _append_preview(replies, result["text"], result["image_url"])

    # ── !video <topic> ────────────────────────────────────────────────────────
    elif text_lower.startswith("!video "):
        topic = text[7:].strip()
        if not topic:
            replies.append({"type": "message",
                             "text": "Example: !video Eid special milk offer 🎬"})
            return replies
        replies.append({"type": "message",
                         "text": f'🎬 Generating HD video ad about "{topic}" (may take ~30s)...'})
        from video_gen import generate_video_ad
        result     = await generate_video_ad(topic)
        post_result = await generate_post(topic)
        set_pending_post(sender_id, {
            "text":      post_result["text"],
            "image_url": result.get("image_url"),
            "video_url": result.get("video_url"),
            "platform":  "both"
        })
        preview = (
            f"🎬 Video Ad Preview:\n\n{post_result['text']}\n\n"
            f"──────────────\n"
            f"Reply yes to post on FB + Instagram ✅\n"
            f"Reply no to cancel ❌"
        )
        replies.append({"type": "message", "text": preview})
        if result.get("image_url"):
            replies.append({"type": "image", "url": result["image_url"]})
        if result.get("video_url"):
            replies.append({"type": "video", "url": result["video_url"]})

    # ── !schedule <topic> | <DD-MM-YYYY> <HH:MM> <fb|ig|both> [post|video] ──
    elif text_lower.startswith("!schedule "):
        parsed = parse_schedule_command(text)
        if not parsed:
            replies.append({"type": "message", "text": (
                "⚠️ Format: !schedule <topic> | <DD-MM-YYYY> <HH:MM> <fb|ig|both> [post|video]\n\n"
                "Example:\n!schedule Eid milk sale | 30-03-2025 10:00 both video"
            )})
            return replies

        replies.append({"type": "message",
                         "text": f'⏳ Preparing scheduled {parsed["post_type"]} for "{parsed["topic"]}"...'})

        if parsed["post_type"] == "video":
            from video_gen import generate_video_ad
            media = await generate_video_ad(parsed["topic"])
            post_res = await generate_post(parsed["topic"])
            post_text  = post_res["text"]
            image_path = media.get("image_url")
            video_path = media.get("video_url")
        else:
            post_res   = await generate_post(parsed["topic"])
            post_text  = post_res["text"]
            image_path = post_res.get("image_url")
            video_path = None

        post_id = create_scheduled_post(
            topic=parsed["topic"], post_text=post_text,
            image_path=image_path, video_path=video_path,
            platform=parsed["platform"], post_type=parsed["post_type"],
            scheduled_time=parsed["scheduled"]
        )
        replies.append({"type": "message", "text": (
            f"✅ Scheduled! Post #{post_id}\n"
            f"📅 {parsed['display_dt']}\n"
            f"📱 Platform: {parsed['platform'].title()}\n"
            f"🎬 Type: {parsed['post_type'].title()}\n\n"
            f"View all: {os.getenv('BASE_URL','http://localhost:8000')}/dashboard"
        )})
        if image_path:
            replies.append({"type": "image", "url": image_path})

    # ── !orders ───────────────────────────────────────────────────────────────
    elif text_lower == "!orders":
        orders = get_all_orders(limit=10)
        if not orders:
            replies.append({"type": "message", "text": "📦 Koi order nahi mila abhi tak."})
        else:
            lines = [f"📦 Total orders: {get_stats()['total']}\n"]
            for o in orders:
                s_emoji = {"pending": "⏳", "confirmed": "✅",
                           "delivered": "🎉", "cancelled": "❌"}.get(o["status"], "📦")
                lines.append(
                    f"{s_emoji} #{o['order_id']} | {o['timestamp']}\n"
                    f"  {o['name']} — {o['product']} — {o['quantity']}\n"
                    f"  📍 {o['address']} | 💳 {o['payment_status']}"
                )
            replies.append({"type": "message", "text": "\n\n".join(lines)})

    # ── !stats ────────────────────────────────────────────────────────────────
    elif text_lower == "!stats":
        s = get_stats()
        replies.append({"type": "message", "text": (
            f"📊 Fresh Go Stats:\n\n"
            f"📦 Total Orders:  {s['total']}\n"
            f"⏳ Pending:       {s['pending']}\n"
            f"✅ Confirmed:     {s['confirmed']}\n"
            f"🎉 Delivered:     {s['delivered']}\n"
            f"💰 Paid:          {s['paid']}\n"
            f"❌ Cancelled:     {s['cancelled']}\n\n"
            f"Dashboard: {os.getenv('BASE_URL','http://localhost:8000')}/dashboard"
        )})

    # ── !update <order_id> <status> ──────────────────────────────────────────
    elif text_lower.startswith("!update "):
        parts = text.split()
        if len(parts) >= 3:
            order_id = parts[1].upper()
            status   = parts[2].lower()
            if status not in ("confirmed", "delivered", "cancelled", "pending"):
                replies.append({"type": "message",
                                 "text": "Status: confirmed / delivered / cancelled / pending"})
                return replies
            order = get_order(order_id)
            if not order:
                replies.append({"type": "message",
                                 "text": f"Order #{order_id} nahi mila."})
                return replies
            update_order_status(order_id, status)
            # Notify customer
            if order.get("customer_id"):
                notify_customer_status(
                    order["customer_id"], order_id, status,
                    order.get("platform", "facebook")
                )
            replies.append({"type": "message",
                             "text": f"✅ Order #{order_id} status → {status.title()}\nCustomer ko notify kar diya."})
        else:
            replies.append({"type": "message",
                             "text": "Format: !update <order_id> <status>\nExample: !update FG0001 confirmed"})

    # ── !help ─────────────────────────────────────────────────────────────────
    elif text_lower == "!help":
        replies.append({"type": "message", "text": (
            "🐄 Fresh Go Admin Commands:\n\n"
            "!post <topic>          →  Image post preview\n"
            "!video <topic>         →  HD video ad preview\n"
            "!schedule <topic> | <DD-MM-YYYY> <HH:MM> <fb|ig|both> [post|video]\n"
            "                       →  Schedule a post/video\n"
            "!orders                →  Last 10 orders\n"
            "!stats                 →  Order statistics\n"
            "!update <id> <status>  →  Update order status\n"
            "!help                  →  This menu\n\n"
            f"Dashboard: {os.getenv('BASE_URL','http://localhost:8000')}/dashboard\n\n"
            "After preview → reply yes/no to publish/cancel."
        )})

    else:
        # Fully conversational admin AI — understands natural language and auto-generates posts
        from ai_engine import admin_chat
        ai_reply, post_topic, analyze_url = await asyncio.to_thread(admin_chat, text)
        replies.append({"type": "message", "text": ai_reply})

        # Auto-analyze Facebook brand page if URL detected
        if analyze_url:
            replies.append({"type": "message",
                             "text": "🔍 Analyzing your Facebook page style... (30-60 seconds)"})
            from brand_analyzer import analyze_facebook_page
            result = await asyncio.to_thread(analyze_facebook_page, analyze_url)
            if result["success"]:
                replies.append({"type": "message", "text": (
                    f"✅ Brand profile saved! Analyzed {result['images_analyzed']} images.\n\n"
                    f"📋 Your brand style:\n{result['profile']}\n\n"
                    "Ab se saari generated images is style ko follow karengi! 🎨"
                )})
            else:
                replies.append({"type": "message", "text": f"⚠️ {result['message']}"})

        # Auto-generate post if AI detected post intent
        if post_topic:
            replies.append({"type": "message",
                             "text": f'⏳ Generating post about "{post_topic}"...'})
            result = await generate_post(post_topic)
            set_pending_post(sender_id, {
                "text": result["text"], "image_url": result["image_url"],
                "video_url": None, "platform": "both"
            })
            _append_preview(replies, result["text"], result["image_url"])

    return replies


async def _do_publish(replies: list, p_text: str, img_url: str,
                       vid_url: str, platform: str):
    """Publish to requested platforms and append result messages."""
    from facebook import publish_video_to_facebook, publish_to_instagram
    import asyncio

    tasks = []
    if platform in ("facebook", "both"):
        if vid_url:
            tasks.append(asyncio.to_thread(publish_video_to_facebook, p_text, vid_url))
        else:
            tasks.append(asyncio.to_thread(publish_post, p_text, img_url))
    if platform in ("instagram", "both"):
        if vid_url:
            tasks.append(asyncio.to_thread(publish_to_instagram, p_text, None, vid_url))
        else:
            tasks.append(asyncio.to_thread(publish_to_instagram, p_text, img_url, None))

    results = await asyncio.gather(*tasks, return_exceptions=True)
    errors  = [r for r in results if isinstance(r, Exception) or
               (isinstance(r, dict) and "error" in r)]

    if errors:
        replies.append({"type": "message",
                         "text": f"⚠️ Some platforms failed: {errors}"})
    else:
        replies.append({"type": "published", "text": p_text,
                         "image_url": img_url, "video_url": vid_url})
        replies.append({"type": "message",
                         "text": f"✅ Published on {platform.title()}! 🎉"})


def _append_preview(replies: list, post_text: str, image_url: str):
    preview = (
        f"📋 Post Preview:\n\n{post_text}\n\n"
        f"──────────────\n"
        f"Reply yes to publish on FB + Instagram ✅\n"
        f"Reply no to cancel ❌"
    )
    replies.append({"type": "message", "text": preview})
    if image_url:
        replies.append({"type": "image", "url": image_url})


# ─────────────────────────────────────────────────────────────────────────────
# Customer flow
# ─────────────────────────────────────────────────────────────────────────────
async def process_customer(sender_id: str, text: str,
                            platform: str = "facebook") -> list:
    if not is_allowed(sender_id):
        wait = seconds_until_reset(sender_id)
        return [{"type": "message", "text": (
            f"Aap ne bohat zyada messages bheje hain 🙏 "
            f"{wait} seconds baad try karen. — Fresh Go 🐄"
        )}]

    if is_history_intent(text):
        return [{"type": "message",
                 "text": get_customer_order_history(sender_id)}]

    if has_active_order(sender_id):
        reply = handle_order_step(sender_id, text)
        return [{"type": "message", "text": reply}]

    if is_order_intent(text):
        reply = start_order(sender_id, platform=platform)
        return [{"type": "message", "text": reply}]

    reply = get_support_reply(text, sender_id, platform=platform)
    return [{"type": "message", "text": reply}]


# ─────────────────────────────────────────────────────────────────────────────
# Facebook Webhook
# ─────────────────────────────────────────────────────────────────────────────
@app.get("/webhook")
async def verify_webhook(
    hub_mode:         str = Query(None, alias="hub.mode"),
    hub_verify_token: str = Query(None, alias="hub.verify_token"),
    hub_challenge:    str = Query(None, alias="hub.challenge"),
):
    if hub_mode == "subscribe" and hub_verify_token == VERIFY_TOKEN:
        print("[Webhook] Verified ✅")
        return int(hub_challenge)
    return JSONResponse(status_code=403, content={"error": "Verification failed"})


@app.post("/webhook")
async def receive_message(request: Request):
    body = await request.json()
    if body.get("object") == "page":
        for entry in body.get("entry", []):
            for event in entry.get("messaging", []):
                sender_id    = event.get("sender", {}).get("id")
                message_data = event.get("message", {})
                text         = message_data.get("text", "").strip()
                if not text or message_data.get("is_echo"):
                    continue
                print(f"[FB MSG] From {sender_id}: {text}")
                if is_admin(sender_id):
                    replies = await process_admin(sender_id, text)
                else:
                    replies = await process_customer(sender_id, text, "facebook")
                for r in replies:
                    if r["type"] == "message":
                        send_message(sender_id, r["text"])
    return {"status": "ok"}


# ─────────────────────────────────────────────────────────────────────────────
# WhatsApp Webhook
# ─────────────────────────────────────────────────────────────────────────────
@app.get("/whatsapp")
async def verify_whatsapp(
    hub_mode:         str = Query(None, alias="hub.mode"),
    hub_verify_token: str = Query(None, alias="hub.verify_token"),
    hub_challenge:    str = Query(None, alias="hub.challenge"),
):
    from fastapi.responses import PlainTextResponse
    print(f"[WhatsApp Verify] mode={hub_mode} token={hub_verify_token} challenge={hub_challenge}")
    expected = os.getenv("WHATSAPP_VERIFY_TOKEN", os.getenv("WEBHOOK_VERIFY_TOKEN", "freshgo_secret_123"))
    if hub_mode == "subscribe" and hub_verify_token == expected and hub_challenge:
        print("[WhatsApp Webhook] Verified ✅")
        return PlainTextResponse(content=str(hub_challenge))
    print(f"[WhatsApp Verify] FAILED — expected token: {expected}")
    return JSONResponse(status_code=403, content={"error": "Verification failed"})


@app.post("/whatsapp")
async def receive_whatsapp(request: Request):
    body = await request.json()
    from whatsapp import parse_whatsapp_webhook, send_whatsapp_message
    messages = parse_whatsapp_webhook(body)
    for msg in messages:
        phone = msg["phone"]
        text  = msg["text"]
        print(f"[WA MSG] From {phone}: {text}")
        replies = await process_customer(phone, text, "whatsapp")
        for r in replies:
            if r["type"] == "message":
                send_whatsapp_message(phone, r["text"])
    return {"status": "ok"}


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard API
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/stats")
async def api_stats():
    return get_stats()


@app.get("/api/orders")
async def api_orders(status: str = Query(None)):
    return get_all_orders(status_filter=status)


@app.post("/api/orders/{order_id}/status")
async def api_update_status(order_id: str, request: Request):
    body         = await request.json()
    new_status   = body.get("status", "").lower()
    pay_status   = body.get("payment_status")
    notes        = body.get("notes")

    valid = {"pending", "confirmed", "delivered", "cancelled"}
    if new_status not in valid:
        return JSONResponse(status_code=400,
                            content={"error": f"Status must be one of {valid}"})

    order = get_order(order_id)
    if not order:
        return JSONResponse(status_code=404, content={"error": "Order not found"})

    update_order_status(order_id, new_status, payment_status=pay_status, notes=notes)

    if order.get("customer_id"):
        notify_customer_status(
            order["customer_id"], order_id, new_status,
            order.get("platform", "facebook")
        )

    return {"ok": True, "order_id": order_id, "status": new_status}


@app.get("/api/scheduled-posts")
async def api_scheduled_posts():
    return get_all_scheduled_posts()


@app.delete("/api/scheduled-posts/{post_id}")
async def api_delete_scheduled(post_id: int):
    deleted = delete_scheduled_post(post_id)
    if not deleted:
        return JSONResponse(status_code=404,
                            content={"error": "Post not found or already published"})
    return {"ok": True}


# ─────────────────────────────────────────────────────────────────────────────
# Brand Assets API
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/brand-assets")
async def api_get_brand_assets():
    from brand_assets import get_asset_urls
    return get_asset_urls()


@app.post("/api/brand-assets/logo")
async def api_upload_logo(file: UploadFile = File(...)):
    from brand_assets import save_logo
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "png"
    data = await file.read()
    url = save_logo(data, ext)
    return {"ok": True, "logo_url": url}


@app.post("/api/brand-assets/packet")
async def api_upload_packet(file: UploadFile = File(...)):
    from brand_assets import save_packet
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "png"
    data = await file.read()
    url = save_packet(data, ext)
    return {"ok": True, "packet_url": url}


# ─────────────────────────────────────────────────────────────────────────────
# Weekly Posts API
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/weekly-post/generate")
async def api_generate_weekly_post():
    """Manually trigger a weekly post generation (for testing or on-demand)."""
    from weekly_posts import generate_weekly_post
    result = await generate_weekly_post()
    return {
        "ok":         True,
        "theme_slug": result["theme_slug"],
        "text":       result["text"],
        "image_url":  result["image_url"],
    }


@app.post("/api/weekly-post/publish-now")
async def api_publish_weekly_now():
    """Generate + immediately publish this week's branded post to Facebook."""
    from weekly_posts import generate_weekly_post
    from facebook import publish_post
    result = await generate_weekly_post()
    await asyncio.to_thread(publish_post, result["text"], result["image_url"])
    return {"ok": True, "theme_slug": result["theme_slug"], "text": result["text"]}


@app.get("/api/weekly-post/themes")
async def api_weekly_themes():
    """Return the full list of 52 themes."""
    from weekly_posts import WEEKLY_THEMES
    from brand_assets import _load
    cfg = _load()
    current_idx = cfg.get("theme_index", 0) % len(WEEKLY_THEMES)
    return {"themes": WEEKLY_THEMES, "next_index": current_idx}


@app.post("/api/schedule-post")
async def api_create_schedule(request: Request):
    body         = await request.json()
    topic        = body.get("topic", "").strip()
    platform     = body.get("platform", "facebook")
    post_type    = body.get("post_type", "image")
    sched_time   = body.get("scheduled_time", "")
    aspect_ratio = body.get("aspect_ratio", "square")
    campaign     = body.get("campaign", "default")

    if not topic or not sched_time:
        return JSONResponse(status_code=400,
                            content={"error": "topic and scheduled_time required"})

    if post_type == "video":
        from video_gen import generate_video_ad
        media      = await generate_video_ad(topic)
        post_res   = await generate_post(topic)
        image_path = media.get("image_url")
        video_path = media.get("video_url")
        post_text  = post_res["text"]
    else:
        from image_gen import fetch_and_save_image
        post_res   = await generate_post(topic)
        post_text  = post_res["text"]
        image_path = await fetch_and_save_image(topic, aspect_ratio, campaign)
        video_path = None

    post_id = create_scheduled_post(
        topic=topic, post_text=post_text, image_path=image_path,
        video_path=video_path, platform=platform, post_type=post_type,
        scheduled_time=sched_time
    )
    return {"ok": True, "post_id": post_id, "preview_text": post_text,
            "image_url": image_path}


@app.post("/api/generate-image")
async def api_generate_image(request: Request):
    """Standalone image generation — no scheduling. Returns image URL immediately."""
    body         = await request.json()
    topic        = body.get("topic", "").strip()
    aspect_ratio = body.get("aspect_ratio", "square")
    campaign     = body.get("campaign", "default")

    from image_gen import fetch_and_save_image, CAMPAIGN_TEMPLATES, ASPECT_RATIOS
    image_url = await fetch_and_save_image(topic, aspect_ratio, campaign)
    if not image_url:
        return JSONResponse(status_code=500, content={"error": "Image generation failed"})
    dims = ASPECT_RATIOS.get(aspect_ratio, ASPECT_RATIOS["square"])
    return {
        "ok": True,
        "image_url":    image_url,
        "aspect_ratio": aspect_ratio,
        "width":        dims["width"],
        "height":       dims["height"],
    }


@app.get("/api/brand/templates")
async def api_brand_templates():
    """Return campaign templates and aspect ratio options for the UI."""
    from image_gen import CAMPAIGN_TEMPLATES, ASPECT_RATIOS
    return {
        "templates":    {k: {"label": v["label"]} for k, v in CAMPAIGN_TEMPLATES.items()},
        "aspect_ratios": ASPECT_RATIOS,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Local Test Chat
# ─────────────────────────────────────────────────────────────────────────────
@app.post("/chat")
async def local_chat(request: Request):
    data       = await request.json()
    message    = data.get("message", "").strip()
    admin_mode = data.get("is_admin", False)
    image_data = data.get("image_data")  # base64 data URL from frontend
    sender_id  = "admin_test" if admin_mode else "customer_test"

    if not message and not image_data:
        return JSONResponse(status_code=400, content={"error": "Message required"})

    # If image attached in admin mode
    if admin_mode and image_data:
        save_keywords = {"brand", "style", "save", "profile", "apni", "meri", "yahi", "isi tarah", "learn"}
        is_save_brand = any(k in (message or "").lower() for k in save_keywords)

        if is_save_brand:
            # Save as brand profile
            from brand_analyzer import analyze_uploaded_images
            result = await asyncio.to_thread(analyze_uploaded_images, [image_data])
            if result["success"]:
                return {"replies": [{"type": "message", "text": (
                    f"✅ Brand profile saved from your uploaded image!\n\n"
                    f"📋 Your brand style:\n{result['profile']}\n\n"
                    "Ab se saari generated images is style ko follow karengi! 🎨"
                )}]}
            else:
                return {"replies": [{"type": "message", "text": "⚠️ Image analyze nahi ho saki. Dobara try karo."}]}
        else:
            # Use as reference for this post only
            from ai_engine import analyze_reference_image
            analysis = await asyncio.to_thread(analyze_reference_image, image_data)
            message = f"[Reference image analysis: {analysis}]\n\nAdmin request: {message or 'isi style mein Fresh Go ki ad banao'}"

    if admin_mode:
        replies = await process_admin(sender_id, message, local_mode=True)
    else:
        replies = await process_customer(sender_id, message)

    return {"replies": replies}


# ─────────────────────────────────────────────────────────────────────────────
# WhatsApp Bulk Messaging
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/wa/status")
async def wa_status():
    return {"mode": "local_script", "configured": True}


@app.post("/api/wa/send-bulk")
async def wa_send_bulk(request: Request):
    return JSONResponse(status_code=503, content={"error": "Use local script"})


@app.get("/api/todays-customers")
async def todays_customers():
    customers = get_todays_customers()
    return {"customers": customers}


@app.get("/api/unpaid-customers")
async def unpaid_customers():
    from database import get_unpaid_customers
    return {"customers": get_unpaid_customers()}


@app.get("/api/export-customers")
async def export_customers():
    """Download today's customers as Excel file."""
    from fastapi.responses import StreamingResponse
    import openpyxl, io
    from datetime import date

    customers = get_todays_customers()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Today's Customers"

    # Header row
    headers = ["Order ID", "Customer Name", "Phone Number", "Product", "Quantity (Litres/KG)", "Address", "Amount (Rs)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=1, column=col).fill = openpyxl.styles.PatternFill("solid", fgColor="D1FAE5")

    # Data rows
    for row, c in enumerate(customers, 2):
        ws.cell(row=row, column=1, value=c.get("order_id", ""))
        ws.cell(row=row, column=2, value=c.get("name", ""))
        ws.cell(row=row, column=3, value=c.get("phone", ""))
        ws.cell(row=row, column=4, value=c.get("product", ""))
        ws.cell(row=row, column=5, value=c.get("quantity", ""))
        ws.cell(row=row, column=6, value=c.get("address", ""))
        ws.cell(row=row, column=7, value=c.get("total_amount", ""))

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"FreshGo_Customers_{date.today().strftime('%d-%m-%Y')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.post("/api/parse-delivery-report")
async def parse_delivery_report(file: UploadFile = File(...)):
    """
    Parse a Rider Delivery Report Excel (columns: #, Date, Customer, Phone, Area,
    Rider, Cash/Credit, Shift, Shop, Type, Order No, Product, Quantity, Rate, Amount, User).
    Returns structured customer list ready for WhatsApp sending.
    """
    import openpyxl, io, re
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
    except Exception as e:
        return JSONResponse(status_code=400, content={"error": f"Cannot read Excel: {e}"})

    ws = wb.active

    # Find header row — skip any title row at the top
    header_row = 1
    for r in range(1, 6):
        cells = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if any("customer" in h or "phone" in h for h in cells):
            header_row = r
            break

    headers = [str(ws.cell(header_row, c).value or "").strip().lower()
               for c in range(1, ws.max_column + 1)]

    def find_col(*keywords):
        for kw in keywords:
            for i, h in enumerate(headers):
                if kw in h:
                    return i
        return None

    c_customer = find_col("customer")
    c_phone    = find_col("phone")
    c_area     = find_col("area")
    c_rider    = find_col("rider")
    c_product  = find_col("product")
    c_qty      = find_col("quantity", "qty")
    c_rate     = find_col("rate")
    c_amount   = find_col("amount")
    c_payment  = find_col("cash")   # "cash/credit" column
    c_date     = find_col("date")

    if c_phone is None:
        return JSONResponse(status_code=400, content={"error": "Phone column not found in Excel"})

    customers = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(v is None for v in row):
            continue
        phone = str(row[c_phone] or "").strip()
        if not phone or phone.lower() == "none" or not any(ch.isdigit() for ch in phone):
            continue

        # Strip trailing (ID) from customer name e.g. "Hassan Muhammad(8)" → "Hassan Muhammad"
        raw_name = str(row[c_customer] or "").strip() if c_customer is not None else "Customer"
        name = re.sub(r"\s*\(\d+\)$", "", raw_name).strip() or "Customer"

        # Normalise phone to Pakistani format
        phone = re.sub(r"[\s\-]", "", phone)
        if phone.startswith("0"):
            phone = "92" + phone[1:]
        elif not phone.startswith("92"):
            phone = "92" + phone

        customers.append({
            "name":     name,
            "phone":    phone,
            "area":     str(row[c_area]    or "").strip() if c_area    is not None else "",
            "rider":    str(row[c_rider]   or "").strip() if c_rider   is not None else "",
            "product":  str(row[c_product] or "Fresh Milk").strip() if c_product is not None else "Fresh Milk",
            "quantity": str(row[c_qty]     or "").strip() if c_qty     is not None else "",
            "rate":     str(row[c_rate]    or "").strip() if c_rate    is not None else "",
            "amount":   str(row[c_amount]  or "").strip() if c_amount  is not None else "",
            "payment":  str(row[c_payment] or "").strip() if c_payment is not None else "",
            "date":     str(row[c_date]    or "").strip() if c_date    is not None else "",
        })

    riders = {}
    for c in customers:
        r = c["rider"] or "Unknown"
        riders[r] = riders.get(r, 0) + 1

    return {"customers": customers, "count": len(customers), "riders": riders}


@app.post("/api/send-delivery-messages")
async def send_delivery_messages(request: Request):
    """
    Accept Excel file upload or JSON list of {name, phone}.
    Send WhatsApp delivery confirmation to each customer.
    """
    from whatsapp import send_whatsapp_message

    content_type = request.headers.get("content-type", "")

    if "multipart/form-data" in content_type:
        from fastapi import UploadFile
        import openpyxl, io
        form = await request.form()
        file = form.get("file")
        custom_msg = form.get("message", "")

        if not file:
            return JSONResponse(status_code=400, content={"error": "No file uploaded"})

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        customers = []
        headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

        name_col   = next((i for i, h in enumerate(headers) if "name" in h), None)
        phone_col  = next((i for i, h in enumerate(headers) if "phone" in h or "number" in h or "mobile" in h), None)

        if phone_col is None:
            return JSONResponse(status_code=400, content={"error": "Phone number column not found in Excel"})

        for row in ws.iter_rows(min_row=2, values_only=True):
            phone = str(row[phone_col] or "").strip()
            name  = str(row[name_col] or "").strip() if name_col is not None else "Customer"
            if phone and phone != "None":
                customers.append({"name": name, "phone": phone})
    else:
        data = await request.json()
        customers = data.get("customers", [])
        custom_msg = data.get("message", "")

    if not customers:
        return JSONResponse(status_code=400, content={"error": "No customers found"})

    # Send messages
    results = {"sent": 0, "failed": 0, "errors": []}
    for customer in customers:
        phone = customer.get("phone", "").strip()
        name  = customer.get("name", "Customer").strip()

        # Clean phone number — ensure it starts with country code
        phone = phone.replace(" ", "").replace("-", "").replace("+", "")
        if phone.startswith("0"):
            phone = "92" + phone[1:]  # Pakistan code

        message = custom_msg or (
            f"Assalam o Alaikum {name}! 🌿\n\n"
            f"Aaj aapka Fresh Go doodh deliver ho gaya hai. "
            f"100% pure, hormone-free cow milk — Nankana Sahib Farm se seedha aapke ghar tak. 🐄\n\n"
            f"Shukriya Fresh Go choose karne ke liye! ❤️\n"
            f"Koi sawaal ho to WhatsApp karen: 0300-3147887"
        )

        result = send_whatsapp_message(phone, message)
        if result:
            results["sent"] += 1
        else:
            results["failed"] += 1
            results["errors"].append(f"{name} ({phone}): failed")

    return {"success": True, **results,
            "summary": f"✅ {results['sent']} messages sent, ❌ {results['failed']} failed"}


# ─────────────────────────────────────────────────────────────────────────────
# UI Routes
# ─────────────────────────────────────────────────────────────────────────────
@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    # Inline SVG cow emoji favicon — no external file needed
    svg = b'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100"><text y=".9em" font-size="90">&#x1F404;</text></svg>'
    from fastapi.responses import Response
    return Response(content=svg, media_type="image/svg+xml")


NO_CACHE = {
    "Cache-Control": "no-cache, no-store, must-revalidate",
    "Pragma": "no-cache",
    "Expires": "0"
}

@app.get("/login", response_class=HTMLResponse)
async def serve_login(request: Request):
    cookie = request.cookies.get(SESSION_COOKIE)
    if cookie and _decode_session_cookie(cookie):
        return RedirectResponse(url="/dashboard", status_code=302)
    with open(os.path.join(BASE_DIR, "login.html")) as f:
        return HTMLResponse(content=f.read(), headers=NO_CACHE)


@app.post("/api/login")
async def api_login(request: Request):
    body     = await request.json()
    username = body.get("username", "").strip().lower()
    password = body.get("password", "").strip()

    role = None
    # Check owner credentials first
    owner_match = (
        hmac.compare_digest(username, ADMIN_USERNAME.lower()) and
        hmac.compare_digest(password, ADMIN_PASSWORD)
    )
    if owner_match:
        role = "owner"
    else:
        db_user = get_admin_user_by_username(username)
        if db_user and db_user["password"] == password:
            role = db_user["role"]

    if role is None:
        return JSONResponse(status_code=401, content={"error": "Wrong username or password"})

    cookie_val = _make_session_cookie(username, role)
    response   = JSONResponse({"ok": True, "role": role})
    response.set_cookie(SESSION_COOKIE, cookie_val,
                        max_age=SESSION_HOURS * 3600,
                        httponly=True, samesite="lax")
    return response


@app.get("/api/logout")
async def api_logout(request: Request):
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie(SESSION_COOKIE)
    return response


@app.get("/api/me")
async def api_me(request: Request):
    user = getattr(request.state, "user", {})
    return {"username": user.get("username", ""), "role": user.get("role", "")}


# ── User Management (owner only) ──────────────────────────────────────────────

def _require_owner(request: Request):
    user = getattr(request.state, "user", {})
    if user.get("role") != "owner":
        raise Exception("owner_only")

@app.get("/api/users")
async def api_list_users(request: Request):
    user = getattr(request.state, "user", {})
    if user.get("role") != "owner":
        return JSONResponse(status_code=403, content={"error": "Owner access required"})
    return {"users": get_all_admin_users()}


@app.post("/api/users")
async def api_create_user(request: Request):
    user = getattr(request.state, "user", {})
    if user.get("role") != "owner":
        return JSONResponse(status_code=403, content={"error": "Owner access required"})
    body      = await request.json()
    username  = body.get("username", "").strip().lower()
    password  = body.get("password", "").strip()
    full_name = body.get("full_name", "").strip()
    role      = body.get("role", "staff")
    if not username or not password:
        return JSONResponse(status_code=400, content={"error": "username and password required"})
    if role not in ("staff", "owner"):
        return JSONResponse(status_code=400, content={"error": "role must be staff or owner"})
    try:
        uid = create_admin_user(username, password, full_name, role)
        return {"ok": True, "id": uid}
    except Exception:
        return JSONResponse(status_code=409, content={"error": "Username already exists"})


@app.delete("/api/users/{user_id}")
async def api_delete_user(user_id: int, request: Request):
    user = getattr(request.state, "user", {})
    if user.get("role") != "owner":
        return JSONResponse(status_code=403, content={"error": "Owner access required"})
    deleted = delete_admin_user(user_id)
    if not deleted:
        return JSONResponse(status_code=404, content={"error": "User not found"})
    return {"ok": True}


@app.put("/api/users/{user_id}")
async def api_update_user(user_id: int, request: Request):
    user = getattr(request.state, "user", {})
    if user.get("role") != "owner":
        return JSONResponse(status_code=403, content={"error": "Owner access required"})
    body      = await request.json()
    full_name = body.get("full_name")
    password  = body.get("password")
    role      = body.get("role")
    update_admin_user(user_id, full_name, password or None, role)
    return {"ok": True}


@app.get("/", response_class=HTMLResponse)
async def serve_chat():
    with open(os.path.join(BASE_DIR, "chat.html")) as f:
        return HTMLResponse(content=f.read(), headers=NO_CACHE)


@app.get("/dashboard", response_class=HTMLResponse)
async def serve_dashboard():
    with open(os.path.join(BASE_DIR, "dashboard.html")) as f:
        return HTMLResponse(content=f.read(), headers=NO_CACHE)


@app.get("/health")
async def health():
    s = get_stats()
    return {
        "status":           "🐄 Fresh Go AI is running!",
        "admin_configured": bool(ADMIN_FB_ID),
        "total_orders":     s["total"],
        "pending_orders":   s["pending"],
    }


if __name__ == "__main__":
    import uvicorn
    print("🐄 Fresh Go AI Agent starting...")
    print("🔗 FB Webhook:  http://localhost:8000/webhook")
    print("📱 WA Webhook:  http://localhost:8000/whatsapp")
    print("📊 Dashboard:   http://localhost:8000/dashboard")
    print("💻 Chat UI:     http://localhost:8000")
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
